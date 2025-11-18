"""
Pilates Excel Database Extraction Script
Parses Pilates_Summary_Spreadsheet_teachingTracker_v10.xlsm
Extracts all movement data, mappings, and teaching knowledge
"""

import openpyxl
import pandas as pd
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any
import logging

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class PilatesExcelExtractor:
    """Extract Pilates domain knowledge from Excel workbook"""

    def __init__(self, excel_path: str):
        """Initialize extractor with Excel file path"""
        self.excel_path = Path(excel_path)
        self.workbook = None
        self.data = {
            'metadata': {
                'source_file': str(self.excel_path.name),
                'extraction_date': datetime.now().isoformat(),
                'version': 'v10'
            },
            'sheets': {},
            'movements': [],
            'muscle_groups': [],
            'sequencing_rules': [],
            'teaching_cues': [],
            'progress_templates': []
        }

    def load_workbook(self):
        """Load Excel workbook with macro support"""
        logger.info(f"Loading workbook: {self.excel_path}")
        try:
            # Load with openpyxl to preserve macros and formulas
            self.workbook = openpyxl.load_workbook(
                self.excel_path,
                keep_vba=True,
                data_only=False  # Keep formulas for analysis
            )
            logger.info(f"Workbook loaded successfully")
            logger.info(f"Sheets found: {self.workbook.sheetnames}")
            return True
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return False

    def analyze_sheet_structure(self):
        """Analyze all sheets and document their structure"""
        logger.info("Analyzing sheet structure...")

        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]

            # Get sheet dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column

            # Get headers (first row)
            headers = []
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(1, col).value
                if cell_value:
                    headers.append(str(cell_value))

            # Sample first few data rows
            sample_data = []
            for row in range(2, min(6, max_row + 1)):
                row_data = []
                for col in range(1, max_col + 1):
                    cell_value = sheet.cell(row, col).value
                    row_data.append(cell_value)
                sample_data.append(row_data)

            self.data['sheets'][sheet_name] = {
                'max_row': max_row,
                'max_col': max_col,
                'headers': headers,
                'sample_data': sample_data[:3]  # First 3 rows
            }

            logger.info(f"Sheet '{sheet_name}': {max_row} rows, {max_col} cols")
            logger.info(f"  Headers: {headers[:5]}...")  # First 5 headers

        return self.data['sheets']

    def extract_movements(self, sheet_name: str = None):
        """Extract movement catalog from appropriate sheet"""
        logger.info("Extracting movement catalog...")

        # Try to find the movements sheet
        possible_names = ['Movements', 'Movement Catalog', 'Exercises', 'Mat Work']

        target_sheet = None
        if sheet_name:
            target_sheet = sheet_name
        else:
            for name in possible_names:
                if name in self.workbook.sheetnames:
                    target_sheet = name
                    break

        if not target_sheet:
            # Use first sheet as fallback
            target_sheet = self.workbook.sheetnames[0]
            logger.warning(f"Movement sheet not found, using: {target_sheet}")

        logger.info(f"Extracting from sheet: {target_sheet}")

        # Use pandas for easier data extraction
        df = pd.read_excel(
            self.excel_path,
            sheet_name=target_sheet,
            engine='openpyxl'
        )

        logger.info(f"Columns found: {df.columns.tolist()}")

        # Convert to list of dictionaries
        movements = []
        for idx, row in df.iterrows():
            # Clean the row data
            movement = {}
            for col in df.columns:
                value = row[col]
                # Skip NaN values
                if pd.notna(value):
                    movement[str(col)] = value

            if movement:  # Only add non-empty rows
                movement['_row_number'] = idx + 2  # Excel row number (1-indexed + header)
                movements.append(movement)

        self.data['movements'] = movements
        logger.info(f"Extracted {len(movements)} movements")

        return movements

    def extract_all_sheets(self):
        """Extract data from all sheets"""
        logger.info("Extracting all sheets as DataFrames...")

        all_sheets = {}

        for sheet_name in self.workbook.sheetnames:
            try:
                df = pd.read_excel(
                    self.excel_path,
                    sheet_name=sheet_name,
                    engine='openpyxl'
                )

                # Convert to list of dictionaries
                sheet_data = []
                for idx, row in df.iterrows():
                    row_dict = {}
                    for col in df.columns:
                        value = row[col]
                        if pd.notna(value):
                            row_dict[str(col)] = value
                    if row_dict:
                        row_dict['_row_number'] = idx + 2
                        sheet_data.append(row_dict)

                all_sheets[sheet_name] = {
                    'columns': df.columns.tolist(),
                    'row_count': len(sheet_data),
                    'data': sheet_data
                }

                logger.info(f"Sheet '{sheet_name}': {len(sheet_data)} rows, {len(df.columns)} columns")

            except Exception as e:
                logger.error(f"Failed to extract sheet '{sheet_name}': {e}")

        return all_sheets

    def validate_data(self, movements: List[Dict]) -> Dict[str, Any]:
        """Validate extracted movement data"""
        logger.info("Validating data...")

        validation_report = {
            'total_movements': len(movements),
            'missing_fields': {},
            'inconsistencies': [],
            'warnings': []
        }

        # Expected fields for movements
        expected_fields = [
            'Movement_Name', 'Difficulty', 'Primary_Muscles',
            'Duration', 'Breathing_Pattern'
        ]

        for field in expected_fields:
            missing_count = sum(1 for m in movements if field not in m)
            if missing_count > 0:
                validation_report['missing_fields'][field] = missing_count

        # Check for duplicate movement names
        movement_names = [m.get('Movement_Name', '') for m in movements]
        duplicates = [name for name in movement_names if movement_names.count(name) > 1]
        if duplicates:
            validation_report['warnings'].append(f"Duplicate movement names: {set(duplicates)}")

        logger.info(f"Validation complete: {validation_report['total_movements']} movements")

        return validation_report

    def save_to_json(self, output_path: str):
        """Save extracted data to JSON file"""
        logger.info(f"Saving to JSON: {output_path}")

        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(self.data, f, indent=2, ensure_ascii=False, default=str)

        logger.info(f"Data saved successfully: {output_file}")

        # Also save a human-readable summary
        summary_path = output_file.parent / f"{output_file.stem}_summary.txt"
        with open(summary_path, 'w') as f:
            f.write(f"Pilates Excel Extraction Summary\n")
            f.write(f"=" * 50 + "\n\n")
            f.write(f"Source: {self.data['metadata']['source_file']}\n")
            f.write(f"Date: {self.data['metadata']['extraction_date']}\n\n")
            f.write(f"Sheets Found: {len(self.data['sheets'])}\n")
            for sheet_name, info in self.data['sheets'].items():
                f.write(f"  - {sheet_name}: {info['max_row']} rows\n")
            f.write(f"\nMovements Extracted: {len(self.data['movements'])}\n")

        logger.info(f"Summary saved: {summary_path}")

        return output_file


def main():
    """Main extraction workflow"""

    # Paths
    excel_path = "/Users/lauraredmond/Documents/Bassline/Admin/7. Product/Design Documentation/Pilates_Summary_Spreadsheet_teachingTracker_v10.xlsm"
    output_path = "/Users/lauraredmond/Documents/Bassline/Projects/MVP2/backend/data/extracted_data.json"

    logger.info("=" * 60)
    logger.info("PILATES EXCEL EXTRACTION - SESSION 2A")
    logger.info("=" * 60)

    # Initialize extractor
    extractor = PilatesExcelExtractor(excel_path)

    # Load workbook
    if not extractor.load_workbook():
        logger.error("Failed to load workbook. Exiting.")
        return False

    # Analyze structure
    sheet_structure = extractor.analyze_sheet_structure()

    # Extract all sheets
    all_sheets_data = extractor.extract_all_sheets()
    extractor.data['all_sheets'] = all_sheets_data

    # Extract movements (from first/main sheet)
    movements = extractor.extract_movements()

    # Validate data
    validation_report = extractor.validate_data(movements)
    extractor.data['validation_report'] = validation_report

    # Save to JSON
    output_file = extractor.save_to_json(output_path)

    logger.info("=" * 60)
    logger.info("EXTRACTION COMPLETE")
    logger.info(f"Output: {output_file}")
    logger.info("=" * 60)

    return True


if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)
